from django.shortcuts import render, get_object_or_404, redirect
from .models import Producto, Proveedor, Pedido, MovimientoInventario
from .forms import ProductoForm, ProveedorForm, PedidoForm, MovimientoForm
from django.db.models import Q
# Página principal
from django.contrib.auth.decorators import login_required
from django.shortcuts import render

@login_required(login_url='login_redirect')
def base(request):
    print("Acceso a base: usuario autenticado") 
    total_productos = Producto.objects.count()
    total_proveedores = Proveedor.objects.count()
    total_pedidos = Pedido.objects.count()
    total_movimientos = MovimientoInventario.objects.count()

    # Pasa los datos a la plantilla
    context = {
        'total_productos': total_productos,
        'total_proveedores': total_proveedores,
        'total_pedidos': total_pedidos,
        'total_movimientos':total_movimientos,
    }
    return render(request, 'base.html', context) # Para verificar acceso

from django.contrib.auth.decorators import login_required
from django.shortcuts import redirect
# Si no está autenticado, redirigir a login
def login_redirect_view(request):
    messages.warning(request, "Debes autenticarte antes de acceder a esta página.")
    return redirect('login') 

# Productos
@login_required
def lista_productos(request):
    query = request.GET.get('q', '')
    productos = Producto.objects.all()

    if query:
        productos = productos.filter(
            Q(codigo__icontains=query) |
            Q(nombre__icontains=query) |
            Q(descripcion__icontains=query) |
            Q(precio__icontains=query) |
            Q(estado_stock__icontains=query)
        )

    return render(request, 'productos.html', {'productos': productos, 'query': query})
@login_required
def detalle_producto(request, producto_id):
    producto = get_object_or_404(Producto, id=producto_id)
    return render(request, 'detalle_producto.html', {'producto': producto})
@login_required
def crear_producto(request):
    if request.method == 'POST':
        form = ProductoForm(request.POST)
        if form.is_valid():
            producto = form.save(commit=False)
            # El código se genera automáticamente en el modelo
            producto.save()
            return redirect('lista_productos')
    else:
        form = ProductoForm()
    return render(request, 'crear_producto.html', {'form': form})
@login_required
def editar_producto(request, producto_id):
    producto = get_object_or_404(Producto, id=producto_id)
    
    if request.method == 'POST':
        form = ProductoForm(request.POST, instance=producto)
        if form.is_valid():
            form.save()
            return redirect('lista_productos')
    else:
        form = ProductoForm(instance=producto)
    
    return render(request, 'editar_producto.html', {'form': form, 'producto': producto})
@login_required
def eliminar_producto(request, producto_id):
    producto = get_object_or_404(Producto, id=producto_id)
    if request.method == 'POST':
        producto.delete()
        return redirect('lista_productos')
    return render(request, 'eliminar_producto.html', {'producto': producto})

@login_required
# Proveedores
def lista_proveedores(request):
    query = request.GET.get('q', '')
    proveedores = Proveedor.objects.all()

    if query:
        proveedores = proveedores.filter(
            Q(nombre__icontains=query) |
            Q(telefono__icontains=query) |
            Q(direccion__icontains=query)
        )

    return render(request, 'proveedores.html', {'proveedores': proveedores, 'query': query})
@login_required
def detalle_proveedor(request, proveedor_id):
    proveedor = get_object_or_404(Proveedor, id=proveedor_id)
    return render(request, 'detalle_proveedor.html', {'proveedor': proveedor})
@login_required
def crear_proveedor(request):
    if request.method == 'POST':
        form = ProveedorForm(request.POST)
        if form.is_valid():
            proveedor = form.save()
            return redirect('lista_proveedores')
    else:
        form = ProveedorForm()

    return render(request, 'crear_proveedor.html', {'form': form})
@login_required
def editar_proveedor(request, proveedor_id):
    proveedor = get_object_or_404(Proveedor, id=proveedor_id)
    
    if request.method == 'POST':
        form = ProveedorForm(request.POST, instance=proveedor)
        if form.is_valid():
            form.save()
            return redirect('lista_proveedores')
    else:
        form = ProveedorForm(instance=proveedor)
    
    return render(request, 'editar_proveedor.html', {'form': form, 'proveedor': proveedor})
@login_required
def eliminar_proveedor(request, proveedor_id):
    proveedor = get_object_or_404(Proveedor, id=proveedor_id)
    if request.method == 'POST':
        proveedor.delete()
        return redirect('lista_proveedores')
    return render(request, 'eliminar_proveedor.html', {'proveedor': proveedor})
@login_required
# Pedidos
def lista_pedidos(request):
    query = request.GET.get('q', '')
    pedidos = Pedido.objects.all()

    if query:
        pedidos = pedidos.filter(
            Q(codigo__icontains=query) |
            Q(fecha_pedido__icontains=query) |
            Q(proveedor__nombre__icontains=query) |
            Q(estado__icontains=query)
        )

    return render(request, 'pedidos.html', {'pedidos': pedidos, 'query': query})
@login_required
def detalle_pedido(request, pedido_id):
    pedido = get_object_or_404(Pedido, id=pedido_id)
    return render(request, 'detalle_pedido.html', {'pedido': pedido})
@login_required
def crear_pedido(request):
    if request.method == 'POST':
        form = PedidoForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('lista_pedidos')
    else:
        form = PedidoForm()
    return render(request, 'crear_pedido.html', {'form': form})
@login_required
def editar_pedido(request, pedido_id):
    pedido = get_object_or_404(Pedido, id=pedido_id)
    if request.method == 'POST':
        form = PedidoForm(request.POST, instance=pedido)
        if form.is_valid():
            form.save()
            return redirect('lista_pedidos')
    else:
        form = PedidoForm(instance=pedido)
    return render(request, 'editar_pedido.html', {'form': form, 'pedido': pedido})
@login_required
def eliminar_pedido(request, pedido_id):
    pedido = get_object_or_404(Pedido, id=pedido_id)
    if request.method == 'POST':
        pedido.delete()
        return redirect('lista_pedidos')
    return render(request, 'eliminar_pedido.html', {'pedido': pedido})

@login_required
# Movimientos
def lista_movimientos(request):
    query = request.GET.get('q', '')
    movimientos = MovimientoInventario.objects.all()

    if query:
        movimientos = movimientos.filter(
            Q(codigo__icontains=query) |
            Q(lote__icontains=query) |
            Q(cantidad__icontains=query) |
            Q(fecha__icontains=query)
        )

    return render(request, 'movimientos.html', {'movimientos': movimientos, 'query': query})
@login_required
def detalle_movimiento(request, movimiento_id):
    movimiento = get_object_or_404(MovimientoInventario, id=movimiento_id)
    return render(request, 'detalle_movimiento.html', {'movimiento': movimiento})
@login_required
def crear_movimiento(request):
    if request.method == 'POST':
        form = MovimientoForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('lista_movimientos')
    else:
        form = MovimientoForm()
    return render(request, 'crear_movimiento.html', {'form': form})
@login_required
def editar_movimiento(request, movimiento_id):
    movimiento = get_object_or_404(MovimientoInventario, id=movimiento_id)
    if request.method == 'POST':
        form = MovimientoForm(request.POST, instance=movimiento)
        if form.is_valid():
            form.save()
            return redirect('lista_movimientos')
    else:
        form = MovimientoForm(instance=movimiento)
    return render(request, 'editar_movimiento.html', {'form': form, 'movimiento': movimiento})
@login_required
def eliminar_movimiento(request, movimiento_id):
    movimiento = get_object_or_404(MovimientoInventario, id=movimiento_id)
    if request.method == 'POST':
        movimiento.delete()
        return redirect('lista_movimientos')
    return render(request, 'eliminar_movimiento.html', {'movimiento': movimiento})

from django.contrib.auth import logout
from django.shortcuts import redirect

def logout_view(request):
    logout(request)
    return redirect('login')


#PDF
from django.http import HttpResponse
from io import BytesIO
from django.template.loader import get_template
from xhtml2pdf import pisa

def export_pdf(request):
    # Obtener los productos de la base de datos
    productos = Producto.objects.all()
    
    # Renderizar el template con los datos
    template_path = 'export_pdf.html'
    context = {'productos': productos}
    response = BytesIO()
    template = get_template(template_path)
    html = template.render(context)
    pisa_status = pisa.CreatePDF(html, dest=response)
    
    if pisa_status.err:
        return HttpResponse('Error al generar PDF')
    
    response.seek(0)
    
    # Configurar el encabezado Content-Disposition para la descarga
    return HttpResponse(response, content_type='application/pdf', headers={'Content-Disposition': 'attachment; filename="productos.pdf"'})


#EXCEL
import openpyxl
from django.http import HttpResponse
from .models import Producto

def export_excel(request):
    # Crear un libro de trabajo y una hoja
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Productos'

    # Agregar encabezados
    headers = ['Código', 'Nombre', 'Descripción', 'Precio', 'Estado de Stock']
    worksheet.append(headers)

    # Obtener los productos
    productos = Producto.objects.all()
    
    for producto in productos:
        worksheet.append([
            producto.codigo,
            producto.nombre,
            producto.descripcion,
            producto.precio,
            producto.estado_stock
        ])

    # Guardar el archivo en un buffer
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=productos.xlsx'
    
    workbook.save(response)
    return response

#pedidopdf
from django.http import HttpResponse
from io import BytesIO
from django.template.loader import get_template
from xhtml2pdf import pisa
from .models import Pedido  # Asegúrate de importar el modelo adecuado

def export_pedidos_pdf(request):
    # Obtener los pedidos de la base de datos
    pedidos = Pedido.objects.all()
    
    # Renderizar el template con los datos
    template_path = 'export_pdf_pedidos.html'  # Asegúrate de que este archivo exista
    context = {'pedidos': pedidos}
    response = BytesIO()
    template = get_template(template_path)
    html = template.render(context)
    pisa_status = pisa.CreatePDF(html, dest=response)
    
    if pisa_status.err:
        return HttpResponse('Error al generar PDF')
    
    response.seek(0)
    
    # Configurar el encabezado Content-Disposition para la descarga
    return HttpResponse(response, content_type='application/pdf', headers={'Content-Disposition': 'attachment; filename="pedidos.pdf"'})

def export_pedidos_excel(request):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Pedidos'
    
    # Encabezados
    headers = ['Código', 'Fecha de Pedido', 'Proveedor', 'Estado']
    worksheet.append(headers)
    
    # Obtener los pedidos
    pedidos = Pedido.objects.all()

    for pedido in pedidos:
        worksheet.append([
            pedido.codigo,
            pedido.fecha_pedido,  # Usar el nombre correcto del campo
            pedido.proveedor.nombre,
            pedido.estado
        ])

    # Guardar el archivo en un buffer
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=pedidos.xlsx'
    
    workbook.save(response)
    return response

#ppp

def export_proveedores_pdf(request):
    # Obtener los proveedores de la base de datos
    proveedores = Proveedor.objects.all()
    
    # Renderizar el template con los datos
    template_path = 'export_pdf_proveedores.html'  # Asegúrate de que este archivo exista
    context = {'proveedores': proveedores}
    response = BytesIO()
    template = get_template(template_path)
    html = template.render(context)
    pisa_status = pisa.CreatePDF(html, dest=response)
    
    if pisa_status.err:
        return HttpResponse('Error al generar PDF')
    
    response.seek(0)
    
    # Configurar el encabezado Content-Disposition para la descarga
    return HttpResponse(response, content_type='application/pdf', headers={'Content-Disposition': 'attachment; filename="proveedores.pdf"'})
def export_proveedores_excel(request):
    # Crear un libro de trabajo y una hoja
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Proveedores'

    # Agregar encabezados
    headers = ['ID', 'Nombre',  'Teléfono']
    worksheet.append(headers)

    # Obtener los proveedores
    proveedores = Proveedor.objects.all()
    
    for proveedor in proveedores:
        worksheet.append([
            proveedor.id,
            proveedor.nombre,
            proveedor.telefono,
        ])

    # Guardar el archivo en un buffer
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=proveedores.xlsx'
    
    workbook.save(response)
    return response

#mmm
# views.py
from django.http import HttpResponse
from django.shortcuts import render
import pandas as pd
from io import BytesIO
from django.template.loader import render_to_string
from xhtml2pdf import pisa

# Ejemplo de vista para exportar a PDF
def export_pdf_movimientos(request):
    movimientos = MovimientoInventario.objects.all()  # Obtén todos los movimientos
    html = render_to_string('movimientos_pdf.html', {'movimientos': movimientos})
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="movimientos.pdf"'
    pisa_status = pisa.CreatePDF(html, dest=response)
    return response

# Ejemplo de vista para exportar a Excel
import pandas as pd
from django.http import HttpResponse
from .models import MovimientoInventario  # Asegúrate de importar tu modelo

def export_excel_movimientos(request):
    # Obtener los movimientos desde la base de datos
    movimientos = MovimientoInventario.objects.all()

    # Convertir los datos a un DataFrame de Pandas
    data = {
        'Código': [movimiento.codigo for movimiento in movimientos],
        'Producto': [movimiento.producto.nombre for movimiento in movimientos],
        'Tipo Movimiento': [movimiento.tipo_movimiento for movimiento in movimientos],
        'Cantidad': [movimiento.cantidad for movimiento in movimientos],
        'Fecha Movimiento': [movimiento.fecha_movimiento.replace(tzinfo=None).strftime('%d/%m/%Y %H:%M') for movimiento in movimientos]
    }

    df = pd.DataFrame(data)

    # Crear la respuesta HTTP para el archivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=movimientos.xlsx'

    # Usar Pandas para exportar el DataFrame a Excel
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Movimientos')

    return response




from django.contrib.auth.models import User
from django.contrib import messages
from django.shortcuts import render, redirect

def registro(request):
    if request.method == 'POST':
        username = request.POST['username']
        email = request.POST['email']
        password = request.POST['password']
        password_confirm = request.POST['password_confirm']  # Confirmar la contraseña

        if password == password_confirm:  # Verificar que las contraseñas coincidan
            try:
                user = User.objects.create_user(username=username, email=email, password=password)
                user.save()
                messages.success(request, 'Registro exitoso. Puedes iniciar sesión ahora.')
                return redirect('login')  # Redirigir a la página de inicio de sesión
            except Exception as e:
                messages.error(request, f'Ocurrió un error: {str(e)}')
        else:
            messages.error(request, 'Las contraseñas no coinciden.')

    return render(request, 'registro.html')  # Asegúrate de tener una plantilla de registro


from django.contrib.auth import authenticate, login
from django.contrib import messages
from django.shortcuts import render, redirect
from django.views.decorators.http import require_POST

def login_view(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect(request.GET.get('next', 'base'))
        else:
            messages.error(request, "Nombre de usuario o contraseña incorrectos.")
    return render(request, 'login.html')

def login_redirect_view(request):
    messages.warning(request, "Debes autenticarte para acceder a esa página.")
    return redirect('login')


